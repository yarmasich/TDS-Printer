"""Walk a ``Labels/`` tree and ingest XLSX rows into the DB.

Expected layout::

    Labels/<project>/<data_hall>/<discipline>/*.xlsx

Project / DataHall / Discipline records are auto-created (with template_id
left empty — admin assigns one later via the web UI).
"""
from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from sqlmodel import Session, delete, select

from .models import DataHall, Discipline, Import, Label, Project

log = logging.getLogger("tds.importer")

# Bundle header marker inside a cell, e.g. "… BUNDLE #2 …", "BDL #3", "BUNDLE#3"
# (case-insensitive, optional spaces). Group 1 is the bundle number. Only used
# for disciplines with bundle_mode on.
_BUNDLE_RE = re.compile(r"(?:BUNDLE|BDL)\s*#\s*(\d+)", re.IGNORECASE)

# Default colour palette copied from the CLI's ``auto_color()``.
DEFAULT_PALETTE: dict[str, str] = {
    "GPU-MS": "FFE699",
    "DEV-EM": "C6E0B4",
    "NVS-NVM": "DDEBF7",
    "NVS-NVB": "F4B084",
    "NVS-KS": "F4B084",
    "AEC ROCE": "F4B084",
    "AEC SISKIN": "F4B084",
    "SIS-T1-T2": "D9D9D9",
}


@dataclass
class ImportStats:
    files: int = 0
    sheets: int = 0
    rows: int = 0
    skipped: int = 0
    disciplines_created: int = 0


def _color_for(name: str) -> str:
    return DEFAULT_PALETTE.get(name.split()[0], "FFFFFF")


def _get_or_create_project(session: Session, name: str) -> Project:
    p = session.exec(select(Project).where(Project.name == name)).first()
    if p:
        return p
    p = Project(name=name)
    session.add(p)
    session.flush()
    return p


def _get_or_create_hall(session: Session, project: Project, name: str) -> DataHall:
    h = session.exec(
        select(DataHall).where(
            DataHall.project_id == project.id, DataHall.name == name
        )
    ).first()
    if h:
        return h
    h = DataHall(project_id=project.id, name=name)
    session.add(h)
    session.flush()
    return h


def _get_or_create_discipline(
    session: Session, hall: DataHall, name: str
) -> tuple[Discipline, bool]:
    d = session.exec(
        select(Discipline).where(
            Discipline.data_hall_id == hall.id, Discipline.name == name
        )
    ).first()
    if d:
        return d, False
    d = Discipline(data_hall_id=hall.id, name=name, color=_color_for(name))
    session.add(d)
    session.flush()
    return d, True


def import_workbook_into_discipline(
    session: Session,
    path: Path,
    discipline: Discipline,
    *,
    display_filename: Optional[str] = None,
) -> ImportStats:
    """Create one ``Import`` row and attach every parsed label to it."""
    stats = ImportStats(files=1)

    imp = Import(
        discipline_id=discipline.id,
        filename=display_filename or path.name,
    )
    session.add(imp)
    session.flush()  # need imp.id for the labels below

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            stats.sheets += 1
            imp.sheets += 1
            ws = wb[sheet_name]
            current_bundle: Optional[str] = None
            for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                left = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
                right = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""

                # In bundle mode a "BUNDLE #N" / "BDL #N" header row is not a
                # label — it starts a new bundle that tags the labels below it.
                if discipline.bundle_mode:
                    m = _BUNDLE_RE.search(left) or _BUNDLE_RE.search(right)
                    if m:
                        current_bundle = m.group(1)
                        stats.skipped += 1
                        imp.skipped += 1
                        continue

                if not left and not right:
                    stats.skipped += 1
                    imp.skipped += 1
                    continue
                session.add(Label(
                    discipline_id=discipline.id,
                    import_id=imp.id,
                    sheet_name=sheet_name,
                    source_file=str(path),
                    row_idx=idx,
                    left_text=left,
                    right_text=right,
                    bundle=current_bundle if discipline.bundle_mode else None,
                ))
                stats.rows += 1
                imp.rows += 1
    finally:
        wb.close()
    session.commit()
    return stats


def scan_labels_dir(
    session: Session, root: Path, *, wipe: bool = False
) -> ImportStats:
    if wipe:
        session.exec(delete(Label))
        session.commit()
        log.info("Label table wiped before scan.")

    total = ImportStats()
    if not root.exists():
        log.warning("Labels root %s does not exist; nothing imported.", root)
        return total

    for project_dir in sorted(p for p in root.iterdir() if p.is_dir()):
        project = _get_or_create_project(session, project_dir.name)
        for hall_dir in sorted(p for p in project_dir.iterdir() if p.is_dir()):
            hall = _get_or_create_hall(session, project, hall_dir.name)
            for discipline_dir in sorted(p for p in hall_dir.iterdir() if p.is_dir()):
                discipline, created = _get_or_create_discipline(
                    session, hall, discipline_dir.name
                )
                if created:
                    total.disciplines_created += 1
                for xlsx in sorted(discipline_dir.glob("*.xlsx")):
                    if xlsx.name.startswith("~$"):
                        continue
                    log.info(
                        "Importing %s/%s/%s/%s",
                        project_dir.name, hall_dir.name, discipline_dir.name, xlsx.name,
                    )
                    s = import_workbook_into_discipline(session, xlsx, discipline)
                    total.files += s.files
                    total.sheets += s.sheets
                    total.rows += s.rows
                    total.skipped += s.skipped

    log.info(
        "Imported %d files, %d sheets, %d rows (%d skipped, %d new disciplines)",
        total.files, total.sheets, total.rows, total.skipped, total.disciplines_created,
    )
    return total
